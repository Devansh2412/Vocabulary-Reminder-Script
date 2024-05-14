import random
import time
import openpyxl
from plyer import notification

# Load the Excel file
workbook = openpyxl.load_workbook('vocab.xlsx')
sheet = workbook.active

# Get the number of rows in the sheet
num_rows = sheet.max_row

# Function to display the notification
def show_notification(word, meaning):
    notification_title = "Word of the Day"
    notification_message = f"{word}: {meaning}"
    notification.notify(
        title=notification_title,
        message=notification_message,
        timeout=10  # Notification timeout in seconds
    )
row_number = random.randint(2, num_rows)

    # Get the word and meaning from the selected row
word = sheet.cell(row=row_number, column=1).value
meaning = sheet.cell(row=row_number, column=2).value

    # Display the notification
show_notification(word, meaning)
